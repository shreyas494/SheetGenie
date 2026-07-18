import os
import sys
import json
import re
import uuid
import subprocess
import traceback
import pandas as pd
import openpyxl
from google import genai
from google.genai import types

def get_excel_metadata(file_path: str) -> dict:
    """
    Extracts metadata from an Excel workbook including sheet names,
    dimensions, column headers, data types, and a preview of data rows.
    """
    if not os.path.exists(file_path):
        return {"error": "File not found"}
        
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        sheets_metadata = []
        
        for name in wb.sheetnames:
            ws = wb[name]
            
            # Simple column parsing using pandas
            try:
                # read first 50 rows to inspect columns/preview
                df = pd.read_excel(file_path, sheet_name=name, nrows=50)
                columns = [str(col) for col in df.columns]
                
                # Sanitize NaN/inf values for JSON compliance
                import math
                raw_preview = df.head(15).to_dict(orient="records")
                preview = []
                for row in raw_preview:
                    clean_row = {}
                    for k, v in row.items():
                        if isinstance(v, float) and (math.isnan(v) or math.isinf(v)):
                            clean_row[k] = None
                        elif pd.isna(v):
                            clean_row[k] = None
                        else:
                            clean_row[k] = v
                    preview.append(clean_row)
            except Exception as e:
                # Fallback if pandas fails (e.g. empty sheet)
                columns = []
                preview = []
                # read first few rows manually
                for r in range(1, min(15, ws.max_row + 1)):
                    row_dict = {}
                    for c in range(1, min(20, ws.max_column + 1)):
                        col_name = ws.cell(row=1, column=c).value or f"Col{c}"
                        if r == 1:
                            columns.append(str(col_name))
                        else:
                            row_dict[str(col_name)] = ws.cell(row=r, column=c).value
                    if r > 1 and any(row_dict.values()):
                        preview.append(row_dict)

            sheets_metadata.append({
                "name": name,
                "columns": columns,
                "rowCount": ws.max_row,
                "columnCount": ws.max_column,
                "preview": preview
            })
            
        return {"sheets": sheets_metadata}
    except Exception as e:
        return {"error": f"Failed to parse metadata: {str(e)}"}

def extract_python_code(response_text: str) -> str:
    """
    Extracts python code blocks from markdown responses.
    """
    pattern = r"```python\s*(.*?)\s*```"
    match = re.search(pattern, response_text, re.DOTALL)
    if match:
        return match.group(1).strip()
        
    # Fallback to the whole text if no code blocks are found
    if "def edit_workbook" in response_text:
        # try to extract from 'def edit_workbook' to the end or next backticks
        start_idx = response_text.find("def edit_workbook")
        code_sub = response_text[start_idx:]
        end_idx = code_sub.find("```")
        if end_idx != -1:
            return code_sub[:end_idx].strip()
        return code_sub.strip()
        
    return response_text.strip()

def execute_generated_code(code_str: str, input_file: str, output_file: str) -> dict:
    """
    Executes the generated Python code inside a separate subprocess running
    under the virtual environment's python.exe. Captures outputs and errors.
    """
    temp_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "temp_runs")
    os.makedirs(temp_dir, exist_ok=True)
    temp_filename = f"run_{uuid.uuid4().hex}.py"
    temp_filepath = os.path.join(temp_dir, temp_filename)
    
    # Wrap code with execution scaffolding
    wrapper_code = f"""# -*- coding: utf-8 -*-
{code_str}

if __name__ == "__main__":
    import json
    import sys
    try:
        # Normalize paths
        input_p = r"{os.path.abspath(input_file)}"
        output_p = r"{os.path.abspath(output_file)}"
        res = edit_workbook(input_p, output_p)
        print("---RESULT_START---")
        print(json.dumps(res))
        print("---RESULT_END---")
    except Exception as e:
        import traceback
        print("---ERROR_START---", file=sys.stderr)
        traceback.print_exc(file=sys.stderr)
        print("---ERROR_END---", file=sys.stderr)
        sys.exit(1)
"""
    
    with open(temp_filepath, "w", encoding="utf-8") as f:
        f.write(wrapper_code)
        
    try:
        # Determine virtual env Python path
        base_dir = os.path.dirname(os.path.abspath(__file__))
        venv_python = os.path.join(base_dir, "venv", "Scripts", "python.exe")
        if not os.path.exists(venv_python):
            # Unix-like path
            venv_python = os.path.join(base_dir, "venv", "bin", "python")
        if not os.path.exists(venv_python):
            # Fallback to standard system python
            venv_python = sys.executable
            
        result = subprocess.run(
            [venv_python, temp_filepath],
            capture_output=True,
            text=True,
            timeout=45
        )
        
        stdout = result.stdout
        stderr = result.stderr
        
        # Cleanup file
        try:
            os.remove(temp_filepath)
        except Exception:
            pass
            
        result_json = None
        if "---RESULT_START---" in stdout:
            parts = stdout.split("---RESULT_START---")
            if len(parts) > 1:
                json_part = parts[1].split("---RESULT_END---")[0].strip()
                try:
                    result_json = json.loads(json_part)
                except Exception:
                    pass
                    
        error_msg = ""
        if "---ERROR_START---" in stderr:
            parts = stderr.split("---ERROR_START---")
            if len(parts) > 1:
                error_msg = parts[1].split("---ERROR_END---")[0].strip()
                
        if result.returncode != 0 and not error_msg:
            error_msg = stderr or stdout or "Unknown script execution error."
            
        # Clean stdout of markers
        clean_stdout = stdout
        if "---RESULT_START---" in clean_stdout:
            parts = clean_stdout.split("---RESULT_START---")
            clean_stdout = parts[0] + parts[1].split("---RESULT_END---")[-1]
        clean_stdout = clean_stdout.strip()
        
        if result.returncode == 0 and result_json:
            return {
                "success": True,
                "message": result_json.get("message", "Workbook edited successfully."),
                "data": result_json.get("data", {}),
                "stdout": clean_stdout,
                "stderr": stderr.strip()
            }
        else:
            return {
                "success": False,
                "error": error_msg or "Execution failed without returning results.",
                "stdout": clean_stdout,
                "stderr": stderr.strip()
            }
            
    except subprocess.TimeoutExpired:
        try:
            os.remove(temp_filepath)
        except Exception:
            pass
        return {
            "success": False,
            "error": "Excel modification timed out after 45 seconds.",
            "stdout": "",
            "stderr": "TimeoutExpired"
        }
    except Exception as e:
        try:
            os.remove(temp_filepath)
        except Exception:
            pass
        return {
            "success": False,
            "error": f"Subprocess running error: {str(e)}",
            "stdout": "",
            "stderr": traceback.format_exc()
        }

def run_agent_loop(
    file_path: str,
    output_path: str,
    prompt: str,
    api_key: str = None,
    chat_history: list = None
) -> dict:
    """
    Main agent execution loop. Generates Python Excel code, runs it,
    captures feedback, and automatically repairs errors up to 3 times.
    """
    # Initialize Google GenAI client
    # google-genai client will look for GEMINI_API_KEY env var if api_key is None
    try:
        if api_key:
            client = genai.Client(api_key=api_key)
        else:
            # Fallback to environment variable or throw if empty
            if not os.environ.get("GEMINI_API_KEY"):
                return {
                    "success": False,
                    "error": "Gemini API key is not configured. Please supply an API key in settings or env.",
                    "logs": []
                }
            client = genai.Client()
    except Exception as e:
        return {
            "success": False,
            "error": f"Failed to initialize Gemini Client: {str(e)}",
            "logs": []
        }

    # Gather Excel structure metadata
    meta = get_excel_metadata(file_path)
    if "error" in meta:
        return {
            "success": False,
            "error": f"Failed to parse Excel file: {meta['error']}",
            "logs": []
        }
        
    meta_json = json.dumps(meta, indent=2)
    history_context = ""
    if chat_history:
        history_context = "\nChat History Context:\n" + "\n".join(
            [f"{msg['role'].upper()}: {msg['content']}" for msg in chat_history[-6:]]
        )

    system_instruction = """You are SheetGenie, an expert Python developer specialized in Excel automation using openpyxl and pandas.
Your job is to write a single Python function that modifies an Excel workbook according to a user request.

RULES:
1. You must write a Python function named `edit_workbook(input_file: str, output_file: str) -> dict`.
2. Do not include any text outside the python code block (wrapped in ```python ... ```).
3. The function must accept `input_file` (path to read the Excel file) and `output_file` (path to save the modified Excel file). CRITICAL: You MUST write the final workbook to the exact `output_file` path parameter (e.g. `wb.save(output_file)`). Do NOT save it to `input_file` or any other path.
4. Use `openpyxl` for structure, layout, styling, and charts. Use `pandas` only for heavy data manipulations. Note that write_excel using pandas will destroy existing formatting. To preserve styles, load the workbook with openpyxl, modify specific cells/rows/columns, and save.
5. The function MUST return a dictionary with keys:
   - "success": bool (True if successful, False if failed)
   - "message": str (description of changes made)
   - "data": dict (optional extra details, like columns added or charts created)
6. Do not include any print statements in the code unless they are useful debug prints. Do not use show() or interactive plot windows.
7. For formulas, write them exactly as Excel formulas, e.g. `ws['C2'] = "=A2*B2"`. Use UPPERCASE formula names like `=SUM()`, `=AVERAGE()`, `=IF()`, `=VLOOKUP()`.
8. When applying styles, load fonts, alignments, borders, and fills:
   `from openpyxl.styles import Font, PatternFill, Alignment, Border, Side`
9. When creating charts, use:
   `from openpyxl.chart import BarChart, LineChart, Reference, Series`
   Always add the chart to the worksheet using `ws.add_chart(chart, "cell_coordinate")`.
10. Ensure columns match up. Column A is 1, B is 2, etc. Row 1 is usually the header row.
"""

    agent_logs = []
    attempts = 0
    max_attempts = 3
    error_feedback = ""
    
    while attempts < max_attempts:
        attempts += 1
        
        # Build prompt
        prompt_content = f"""Here is the Excel workbook metadata:
{meta_json}
{history_context}

User Request: {prompt}
"""
        if error_feedback:
            prompt_content += f"""
Previous Python code execution failed. Please correct the code and address the error.
Error details:
{error_feedback}

Make sure to fix this specific issue and return a fully working `edit_workbook` function.
"""

        agent_logs.append({
            "attempt": attempts,
            "status": "Generating code...",
            "message": f"Sending request to Gemini (Attempt {attempts}/{max_attempts})..."
        })
        
        try:
            response = client.models.generate_content(
                model='gemini-3.5-flash',
                contents=prompt_content,
                config=types.GenerateContentConfig(
                    system_instruction=system_instruction,
                    temperature=0.1
                )
            )
            
            code = extract_python_code(response.text)
            
            agent_logs[-1]["code"] = code
            agent_logs[-1]["status"] = "Running code..."
            
            # Execute
            exec_res = execute_generated_code(code, file_path, output_path)
            
            agent_logs[-1]["stdout"] = exec_res.get("stdout", "")
            agent_logs[-1]["stderr"] = exec_res.get("stderr", "")
            
            if exec_res["success"]:
                # Verify that the output file was actually written to output_path!
                if not os.path.exists(output_path):
                    exec_res["success"] = False
                    exec_res["error"] = f"The script returned success, but failed to write the output file to the specified path: {output_path}. Make sure you write the workbook using `wb.save(output_file)` at the end of your function!"
                
            if exec_res["success"]:
                agent_logs[-1]["status"] = "Success"
                agent_logs[-1]["message"] = exec_res["message"]
                
                # Retrieve updated metadata
                updated_meta = get_excel_metadata(output_path)
                
                return {
                    "success": True,
                    "message": exec_res["message"],
                    "data": exec_res.get("data", {}),
                    "code": code,
                    "logs": agent_logs,
                    "metadata": updated_meta
                }
            else:
                agent_logs[-1]["status"] = "Failed"
                agent_logs[-1]["error"] = exec_res["error"]
                error_feedback = f"Code raised an error:\n{exec_res['error']}\n\nGenerated Code was:\n{code}"
                
        except Exception as e:
            agent_logs[-1]["status"] = "API Error"
            agent_logs[-1]["error"] = str(e)
            error_feedback = f"Gemini API or internal error occurred: {str(e)}"
            
    return {
        "success": False,
        "error": "Failed to modify Excel sheet after multiple attempts.",
        "logs": agent_logs
    }

def convert_pdf_to_xlsx(pdf_path: str, xlsx_path: str, api_key: str = None) -> dict:
    """
    Uploads a PDF to Gemini, extracts tabular structures, and generates
    a formatted Excel sheet locally using a sandbox code-execution loop.
    """
    try:
        if api_key:
            client = genai.Client(api_key=api_key)
        else:
            if not os.environ.get("GEMINI_API_KEY"):
                return {"success": False, "error": "Gemini API key is not configured."}
            client = genai.Client()
    except Exception as e:
        return {"success": False, "error": f"Failed to initialize Gemini Client: {str(e)}"}
        
    try:
        # 1. Upload the PDF file to Gemini Files API
        gemini_file = client.files.upload(file=open(pdf_path, 'rb'), mime_type="application/pdf")
        
        # 2. Build prompt and instructions
        prompt = "Read this PDF, extract all tabular data, tables, and lists, and generate a Python openpyxl script to write it to Excel."
        system_instruction = """You are a precise data extractor. Your goal is to inspect the uploaded PDF document, extract all tables, data rows, and structural relationships, and write a Python script that creates an Excel workbook.

RULES:
1. You must write a Python function named `create_workbook_from_pdf(output_file: str)`.
2. Do not include any text outside the python code block (wrapped in ```python ... ```).
3. The function must accept `output_file` (the path where the .xlsx file will be saved).
4. Extract the actual data values from the PDF and hardcode them as Python lists of rows/dictionaries inside the generated script. Do NOT try to read the PDF locally in the python script. Translate the text/tables you read into data structures in your code.
5. Create worksheets for distinct tables or sections. Design the layout using openpyxl: set columns, style headers (bold font, gray or blue fills, centered alignment), draw light cell borders, and set number formatting (dates, currency, percentage).
"""

        # 3. Call Gemini 3.5 Flash
        response = client.models.generate_content(
            model='gemini-3.5-flash',
            contents=[gemini_file, prompt],
            config=types.GenerateContentConfig(
                system_instruction=system_instruction,
                temperature=0.1
            )
        )
        
        # Clean up cloud file
        try:
            client.files.delete(name=gemini_file.name)
        except Exception as delete_err:
            print(f"Failed to delete Gemini file: {delete_err}")
            
        code = extract_python_code(response.text)
        
        # 4. Run the code in a subprocess to generate the excel file
        temp_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "temp_runs")
        os.makedirs(temp_dir, exist_ok=True)
        temp_filename = f"pdf_run_{uuid.uuid4().hex}.py"
        temp_filepath = os.path.join(temp_dir, temp_filename)
        
        wrapper_code = f"""# -*- coding: utf-8 -*-
{code}

if __name__ == "__main__":
    import sys
    try:
        create_workbook_from_pdf(r"{os.path.abspath(xlsx_path)}")
        print("---PDF_SUCCESS---")
    except Exception as e:
        import traceback
        traceback.print_exc(file=sys.stderr)
        sys.exit(1)
"""
        with open(temp_filepath, "w", encoding="utf-8") as f:
            f.write(wrapper_code)
            
        # Determine python path
        base_dir = os.path.dirname(os.path.abspath(__file__))
        venv_python = os.path.join(base_dir, "venv", "Scripts", "python.exe")
        if not os.path.exists(venv_python):
            venv_python = os.path.join(base_dir, "venv", "bin", "python")
        if not os.path.exists(venv_python):
            venv_python = sys.executable
            
        result = subprocess.run(
            [venv_python, temp_filepath],
            capture_output=True,
            text=True,
            timeout=60
        )
        
        # Cleanup temp file
        try:
            os.remove(temp_filepath)
        except:
            pass
            
        if result.returncode == 0 and "---PDF_SUCCESS---" in result.stdout:
            return {
                "success": True,
                "message": "Successfully extracted PDF data into Excel.",
                "code": code
            }
        else:
            error_details = result.stderr or result.stdout or "Subprocess failed without output."
            return {
                "success": False,
                "error": f"Failed to execute data extraction script:\n{error_details}",
                "code": code
            }
            
    except Exception as e:
        return {
            "success": False,
            "error": f"PDF conversion loop failed: {str(e)}",
            "code": ""
        }

def excel_to_pdf_win32(excel_path: str, pdf_path: str, active_sheet_name: str = None) -> bool:
    """
    Converts Excel sheet to PDF using native COM automation (pywin32) on Windows.
    """
    import win32com.client
    import pythoncom
    
    pythoncom.CoInitialize()
    excel_app = None
    try:
        excel_app = win32com.client.Dispatch("Excel.Application")
        excel_app.Visible = False
        excel_app.ScreenUpdating = False
        excel_app.DisplayAlerts = False
        
        abs_excel = os.path.abspath(excel_path)
        abs_pdf = os.path.abspath(pdf_path)
        
        wb = excel_app.Workbooks.Open(abs_excel)
        
        # Select active sheet if name is specified
        if active_sheet_name:
            try:
                ws = wb.Sheets(active_sheet_name)
                ws.Select()
            except Exception:
                pass # Fallback to active sheet
                
        # 0 corresponds to xlTypePDF
        wb.ActiveSheet.ExportAsFixedFormat(0, abs_pdf)
        wb.Close(SaveChanges=False)
        return True
    except Exception as e:
        print(f"Win32 Excel PDF conversion error: {str(e)}")
        raise e
    finally:
        if excel_app:
            try:
                excel_app.Quit()
            except:
                pass
        pythoncom.CoUninitialize()

def excel_to_pdf_fallback(excel_path: str, pdf_path: str, active_sheet_name: str = None) -> bool:
    """
    Converts Excel sheet to PDF using ReportLab as a lightweight fallback.
    """
    from reportlab.lib.pagesizes import letter, landscape
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors
    
    try:
        # Load the sheet using openpyxl
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        if active_sheet_name and active_sheet_name in wb.sheetnames:
            ws = wb[active_sheet_name]
        else:
            ws = wb.active
            
        # Extract all values to a list of lists
        data = []
        max_c = min(15, ws.max_column + 1)
        max_r = min(100, ws.max_row + 1)
        
        for r in range(1, max_r):
            row_vals = []
            for c in range(1, max_c):
                val = ws.cell(row=r, column=c).value
                row_vals.append(str(val) if val is not None else "")
            if any(row_vals):
                data.append(row_vals)
                
        if not data:
            data = [["Sheet is Empty"]]
            
        # Setup ReportLab Document in Landscape Letter format
        doc = SimpleDocTemplate(pdf_path, pagesize=landscape(letter), rightMargin=20, leftMargin=20, topMargin=20, bottomMargin=20)
        elements = []
        
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'TitleStyle',
            parent=styles['Heading1'],
            fontSize=16,
            textColor=colors.HexColor('#111827'),
            spaceAfter=15
        )
        
        # Add Sheet Title
        elements.append(Paragraph(f"Spreadsheet Report: {ws.title}", title_style))
        elements.append(Spacer(1, 10))
        
        # Create Table Flowable
        table = Table(data)
        
        # Table Styling
        t_style = TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3B82F6')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('TOPPADDING', (0, 0), (-1, 0), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E5E7EB')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F9FAFB')]),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
            ('TOPPADDING', (0, 1), (-1, -1), 6),
        ])
        
        table.setStyle(t_style)
        elements.append(table)
        
        doc.build(elements)
        return True
    except Exception as e:
        print(f"Reportlab PDF fallback creation error: {str(e)}")
        raise e

def convert_excel_to_pdf(excel_path: str, pdf_path: str, active_sheet_name: str = None) -> bool:
    """
    Tries win32 COM conversion first, falls back to reportlab if Excel is not installed.
    """
    try:
        import win32com.client
        return excel_to_pdf_win32(excel_path, pdf_path, active_sheet_name)
    except Exception as e:
        print("Win32 PDF export failed or unavailable. Falling back to ReportLab...")
        return excel_to_pdf_fallback(excel_path, pdf_path, active_sheet_name)

